from flask import request, jsonify, Flask, send_file, session

import uuid

from .. import db
from .models import Account

import os
import pika
from flask_caching import Cache
import redis

import xlsxwriter

from openpyxl import load_workbook

import jwt, os
from datetime import datetime

from flask import request, abort

from functools import wraps

from cryptography.fernet import Fernet



# ----------------------------------------------- #

# Query Object Methods => https://docs.sqlalchemy.org/en/14/orm/query.html#sqlalchemy.orm.Query
# Session Object Methods => https://docs.sqlalchemy.org/en/14/orm/session_api.html#sqlalchemy.orm.Session
# How to serialize SqlAlchemy PostgreSQL Query to JSON => https://stackoverflow.com/a/46180522

app = Flask(__name__)
app.config["CACHE_TYPE"] = os.getenv("CACHE_TYPE")
app.config["CACHE_REDIS_HOST"] = os.getenv("CACHE_REDIS_HOST")
app.config["CACHE_REDIS_PORT"] = os.getenv("CACHE_REDIS_PORT")
app.config["CACHE_REDIS_DB"] = os.getenv("CACHE_REDIS_DB")

cache = Cache(app=app)
cache.init_app(app)

# Initialize Redis client
redis_client = redis.Redis(
    host=os.getenv("CACHE_REDIS_HOST"), port=os.getenv("CACHE_REDIS_PORT"), db=0
)


encryption_key = Fernet.generate_key()
fernet = Fernet(encryption_key)

def encrypt_data(data):
    encrypted_data = fernet.encrypt(data.encode())
    return encrypted_data

def decrypt_data(encrypted_data):
    decrypted_data = fernet.decrypt(encrypted_data).decode()
    return decrypted_data

def ref_token():
    try:
            data = jwt.decode(decrypt_data(session['refresh_token']), os.getenv("SECRET_REFRESH_TOKEN"), algorithms=["HS256"])
            current_user = Account().query.filter_by(email=data["email"]).first()
            if current_user.email is None:
                return {
                    "message": "Invalid Authentication token!",
                    "error": "Unauthorized",
                }, 401
            if not current_user.email:
                abort(403)
            userd = jwt.encode(
                    {"email": current_user.email},
                    os.getenv("SECRET_KEY"),
                    algorithm="HS256",
                )

            return {"message": "Successfully fetched new auth token", "new token": userd}        
    except Exception as e:
            return {
                "message": "Invalid Authentication token!",
                "error": "Unauthorized",
            }, 500 

def token_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        token = None
        if "Authorization" in request.headers:
            token = request.headers["Authorization"].split(" ")[1]
        if not token:
            return {
                "message": "Authentication Token is missing!",
                "error": "Unauthorized",
            }, 401
        try:
            data = jwt.decode(token, os.getenv("SECRET_KEY"), algorithms=["HS256"])
            current_user = Account().query.filter_by(email=data["email"]).first()
            if current_user.email is None:
                return {
                    "message": "Invalid Authentication token!",
                    "error": "Unauthorized",
                }, 401
            if not current_user.email:
                abort(403)
        except Exception as e:
            return {
                "message": "Something went wrong",
                "data": None,
                "error": str(e),
            }, 500

        return f(*args, **kwargs)

    return wrapper


def login():
    try:
        data = request.json
        if not data:
            return {
                "message": "Please provide user details",
                "data": None,
                "error": "Bad request",
            }, 400

        user = (
            Account()
            .query.filter_by(email=data["email"])
            .filter_by(password=data["password"])
            .first()
        )
        print(111, user.email)
        if user:
            try:

                # token should expire after 24 hrs
                userd = jwt.encode(
                    {"email": user.email},
                    os.getenv("SECRET_KEY"),
                    algorithm="HS256",
                )

                refresh_token = jwt.encode(
                    {"email": user.email},
                    os.getenv("SECRET_REFRESH_TOKEN"),
                    algorithm="HS256",
                )
                encrypted_data = encrypt_data(refresh_token)
                print(111, refresh_token)
                session["refresh_token"] = encrypted_data
                return {"message": "Successfully fetched auth token", "token": userd}
            except Exception as e:
                return {"error": "Something went wrong", "message": str(e)}, 500
        return {
            "message": "Error fetching auth token!, invalid email or password",
            "data": None,
            "error": "Unauthorized",
        }, 404
    except Exception as e:
        return {"message": "Something went wrong!", "error": str(e), "data": None}, 500


def list_all_accounts_controller():
    accounts = Account.query.all()
    response = []

    for account in accounts:
        response.append(account.toDict())
    redis_client.set(str(uuid.uuid4()), repr(response))
    return jsonify(response)


def read_report_controller():
    excelfile = "fileUpload/report.xlsx"
    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    data = {}
    connection = pika.BlockingConnection(pika.ConnectionParameters(host="localhost"))
    channel = connection.channel()

    channel.queue_declare(queue="send.email.flask")
    for row in range(2, ws.max_row + 1):  # need +1 to get last row!
        for col in "A":  # A gets players for texted season
            cell_name = "{}{}".format(col, row)
            data["email"] = ws[cell_name].value
        for col in "B":  # A gets players for texted season
            cell_name = "{}{}".format(col, row)
            data["username"] = ws[cell_name].value
        for col in "C":  # A gets players for texted season
            cell_name = "{}{}".format(col, row)
            data["dob"] = ws[cell_name].value
        for col in "D":  # A gets players for texted season
            cell_name = "{}{}".format(col, row)
            data["country"] = ws[cell_name].value
        for col in "E":  # A gets players for texted season
            cell_name = "{}{}".format(col, row)
            data["phone_number"] = ws[cell_name].value
        print(12122, data)
        if data["email"] is not None:
            channel.basic_publish(
                exchange="save.data.and.send.email.flask",
                routing_key="save.data.and.send.email.flask",
                body=repr(data),
            )
    return jsonify(data)


def createReport(data):
    now = datetime.now()
    fn = "generated/fileCustomer" + now.strftime("%d%m%Y%H%M%S") + ".xlsx"
    f = open(fn, "x")
    workbook = xlsxwriter.Workbook(fn)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, "No: ")
    worksheet.write(0, 1, "Email: ")
    worksheet.write(0, 2, "Username: ")
    worksheet.write(0, 3, "DOB: ")
    worksheet.write(0, 4, "Country: ")
    worksheet.write(0, 5, "Phone Number: ")
    row = 0
    for k in data:
        col = 0
        row += 1
        worksheet.write(row, col, row)
        col += 1
        worksheet.write(row, col, k["email"])
        col += 1
        worksheet.write_string(row, col, k["username"])
        col += 1
        worksheet.write_string(row, col, str(k["dob"]))
        col += 1
        worksheet.write_string(row, col, k["country"])
        col += 1
        worksheet.write_string(row, col, k["phone_number"])

    workbook.close()
    return fn


def export_all_accounts_excel_controller():

    accounts = Account.query.all()
    response = []
    for account in accounts:
        response.append(account.toDict())

    fn = createReport(response)
    return send_file(fn, mimetype="application/vnd.ms-excel")


# return jsonify(response)


def create_account_controller():
    print(111, request.get_json())
    # request_form = request.form.to_dict()

    data = request.get_json()

    connection = pika.BlockingConnection(pika.ConnectionParameters(host="localhost"))
    channel = connection.channel()

    channel.queue_declare(queue="save.data.and.send.email.flask")

    channel.basic_publish(
        exchange="", routing_key="save.data.and.send.email.flask", body=repr(data)
    )
    print(data)

    # maildata(data)
    return jsonify("success")


def retrieve_account_controller(account_id):
    response = Account.query.get(account_id).toDict()
    return jsonify(response)


def update_account_controller(account_id):
    request_form = request.get_json()
    request_form["id"] = account_id

    connection = pika.BlockingConnection(pika.ConnectionParameters(host="localhost"))
    channel = connection.channel()

    channel.queue_declare(queue="update.data")

    channel.basic_publish(
        exchange="", routing_key="update.data", body=repr(request_form)
    )

    return jsonify(request_form)


def delete_account_controller(account_id):
    Account.query.filter_by(id=account_id).delete()
    db.session.commit()

    return ('Account with Id "{}" deleted successfully!').format(account_id)
